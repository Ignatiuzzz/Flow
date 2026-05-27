from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


EXCEL_EXPORT_DIR = Path("app/uploads/exports/excel")


def ensure_excel_directory() -> None:
    EXCEL_EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def safe_text(value: object) -> str:
    if value is None:
        return ""

    return str(value)


def build_excel_filename(project_id: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"matriz_proyecto_{project_id}_{timestamp}.xlsx"


def generate_matrix_excel(project_id: str, findings: List[dict]) -> str:
    ensure_excel_directory()

    filename = build_excel_filename(project_id)
    file_path = EXCEL_EXPORT_DIR / filename

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matriz de Hallazgos"

    headers = [
        "Código",
        "Nombre",
        "Descripción",
        "Criterio",
        "Objetivo",
        "Causa",
        "Efecto",
        "Conclusión",
        "Impacto",
        "Urgencia",
        "Riesgo",
        "Nivel",
        "Recomendaciones",
        "Fecha de Creación",
        "Fecha de Actualización",
    ]

    sheet.append(headers)

    for finding in findings:
        sheet.append([
            safe_text(finding.get("codigo")),
            safe_text(finding.get("nombre")),
            safe_text(finding.get("descripcion")),
            safe_text(finding.get("criterio")),
            safe_text(finding.get("objetivo")),
            safe_text(finding.get("causa")),
            safe_text(finding.get("efecto")),
            safe_text(finding.get("conclusion")),
            finding.get("impacto", ""),
            finding.get("urgencia", ""),
            finding.get("riesgo", ""),
            safe_text(finding.get("nivel")),
            safe_text(finding.get("recomendaciones")),
            safe_text(finding.get("fechaCreacion")),
            safe_text(finding.get("fechaActualizacion")),
        ])

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    column_widths = {
        "A": 15,
        "B": 30,
        "C": 45,
        "D": 45,
        "E": 45,
        "F": 35,
        "G": 35,
        "H": 35,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 15,
        "M": 45,
        "N": 25,
        "O": 25,
    }

    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    workbook.save(file_path)

    return str(file_path).replace("\\", "/")